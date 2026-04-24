import openpyxl


def find_column_number(file_path, target_header):
    """
    在指定的 Excel 文件中查找特定列名所在的列序号。

    :param file_path: Excel 文件的路径
    :param target_header: 要查找的列名（表头）
    :return: 列序号 (int)，如果未找到则返回 None
    """
    try:
        # 1. 加载工作簿
        workbook = openpyxl.load_workbook(file_path)
        # 获取当前活动的工作表
        sheet = workbook.active

        # 2. 遍历第一行的所有列
        # sheet.max_column 获取表格的最大列数
        for col_index in range(1, sheet.max_column + 1):
            # 获取第一行当前列单元格的值
            header_value = sheet.cell(row=1, column=col_index).value

            # 3. 判断是否匹配
            if header_value == target_header:
                return col_index

        # 如果循环结束仍未找到，返回 None
        print(f"未找到列名：'{target_header}'")
        return None

    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'，请检查路径是否正确。")
        return None
    except Exception as e:
        print(f"发生未知错误：{e}")
        return None


# --- 使用示例 ---
excel_file = r'C:\Users\Administrator\Desktop\upgrade_rate.xlsx'  # 替换为你的实际文件路径
column_name = '未升级次数(0)'

result = find_column_number(excel_file, column_name)

if result:
    print(f"列名 '{column_name}' 位于第 {result} 列。")
else:
    print("查找失败。")